###フェイシャルシートの質問に重複があるのを修正###

import os
import sys
import openpyxl as excel
import shutil

#学習用データの保存ディレクトリ
dare = '重複なしデータ'
base_dir = 'S:/個人作業用/宇野/ワールドジャパン/学習用データ(テスト用)/' + dare + '/【A】/【要修正】'

#ディレクトリ内の全てを取得
files_dirs = os.listdir(base_dir)
#ファイルのみを取得
files = [f for f in files_dirs if os.path.isfile(os.path.join(base_dir, f))]
#ファイル数を取得
len_files = str(len(files))

#ファイルが存在しなかった場合は終了
if int(len_files) == 0:
  sys.exit()

#プログレスバーの表示に使用
j = 0
bar_size = 20

#修正済み確認リスト用の配列
ary = []

#各ファイルをチェック
for nam in files:
    j = j + 1
    #なぜか取得したファイル名の最初に~$が入っていることがあるため、それを除去
    #取得したファイル名とベースディレクトリを結合
    if '~$' in nam :
        path = base_dir + '/' + nam[2:]
        #print(path)
    else :
        path = base_dir + '/' + nam
        #print(path)

    #ファイルの拡張子をチェック
    name,ext = os.path.splitext(path)
    if ext != '.xlsx' :
        continue
    
    #プログレスバーの表示
    rate = round(j / int(len_files) * bar_size) - 1
    pro_bar =('=' * rate) + (' ' * (bar_size - rate))
    print('\r' + '[' + pro_bar + ']' + ' ' + str(round((j + 1) / int(len_files) * 100, 2)) + '% ' + nam + '             ', end='')

    #ファイルを開く
    bk = excel.load_workbook(path)
    sheet = bk.worksheets[4]

    #項目数をカウント
    if sheet.cell(row = 41, column = 3).value == 'アトピー' and sheet.cell(row = 42, column = 3).value == 'アレルギー' and sheet.cell(row = 43, column = 3).value == 'アトピー' and sheet.cell(row = 44, column = 3).value == 'アレルギー':
        if sheet.cell(row = 43, column = 4).value != None or sheet.cell(row = 44, column = 4).value != None :
            sheet.cell(row = 41, column = 4).value = sheet.cell(row = 43, column = 4).value
            sheet.cell(row = 41, column = 4).value = sheet.cell(row = 43, column = 4).value
        
        sheet.delete_rows(44)
        sheet.cell(row = 43, column = 3).value = 'その他'
        sheet.cell(row = 43, column = 4).value = ''

        bk.save(path)

        bk = excel.load_workbook(path)

    k = 0

    for sheet in bk :
        i = 1
        val = 0
        if 'ボディ' in sheet.title :
            #項目数をカウント
            while val != '契約内容' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しい項目数の場合は修正済み確認リストに追加
            if i == 131 :
                k += 1
        elif 'バスト' in sheet.title :
            #項目数をカウント
            while val != '契約内容' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しい項目数の場合は修正済み確認リストに追加
            if i == 95 :
                k += 1
        elif 'フェイシャル' in sheet.title :
            #項目数をカウント
            while val != '契約内容' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しい項目数の場合は修正済み確認リストに追加
            if i == 72 :
                k += 1
        elif '脱毛' in sheet.title :
            #項目数をカウント
            while val != '医師から注意を受けている事や体質的に気になる事' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しい項目数の場合は修正済み確認リストに追加
            if i == 62 :
                k += 1
    #print('\n' + str(k))
    if k == 4 :
        ary.append(path)


#プログレスバーの表示
rate = bar_size
pro_bar =('=' * rate) + (' ' * (bar_size - rate))
print('\r' + '[' + pro_bar + ']' + ' ' + str(round(j / int(len_files) * 100, 2)) + '%', end='')

ary = list(set(ary))

#例外ファイルを要修正フォルダに移動させる
for one in ary :
    shutil.move(one, one.replace('/【A】/【要修正】/', '/【A】/'))

print('\n合計数 : ' + str(len(ary)))
print('Complete!',end = '')
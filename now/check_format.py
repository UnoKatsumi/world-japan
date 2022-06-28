###フォーマットが正しくないファイルをリストアップ###

import os
import openpyxl as excel
import shutil

#学習用データの保存ディレクトリ
dare = '重複なしデータ'
base_dir = 'S:/個人作業用/渡邊/ワールドジャパン/' + dare + '/【A】'

#ディレクトリ内の全てを取得
files_dirs = os.listdir(base_dir)
#ファイルのみを取得
files = [f for f in files_dirs if os.path.isfile(os.path.join(base_dir, f))]
#ファイル数を取得
len_files = str(len(files))

#プログレスバーの表示に使用
j = 0
bar_size = 20

#例外リスト用の配列
ary = []

#各ファイルをチェック
for nam in files:
    j = j + 1
    #なぜか取得したファイル名の最初に~$が入っていることがあるため、それを除去
    #取得したファイル名とベースディレクトリを結合
    if '~$' in nam :
        path = base_dir + '/' + nam[2:]
        #print(path)
    else :
        path = base_dir + '/' + nam
        #print(path)

    #ファイルの拡張子をチェック
    name,ext = os.path.splitext(path)
    if ext != '.xlsx' :
        continue
    
    #プログレスバーの表示
    rate = round(j / int(len_files) * bar_size) - 1
    pro_bar =('=' * rate) + (' ' * (bar_size - rate))
    print('\r' + '[' + pro_bar + ']' + ' ' + str(round((j + 1) / int(len_files) * 100, 2)) + '% ' + nam + '             ', end='')

    #ファイルを開く
    bk = excel.load_workbook(path)

    #ファイル内の各シートを読込み
    for sheet in bk :
        i = 1
        val = 0
        if 'ボディ' in sheet.title :
            #項目数をカウント
            while val != '契約内容' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しくない項目数の場合は例外ファイルリストに追加
            if i != 131 :
                ary.append(path)
                print('\n' + path + '\n' + str(sheet.title) + str(i))
                break
        elif 'バスト' in sheet.title :
            #項目数をカウント
            while val != '契約内容' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しくない項目数の場合は例外ファイルリストに追加
            if i != 95 :
                ary.append(path)
                print('\n' + path + '\n' + str(sheet.title) + str(i))
                break
        elif 'フェイシャル' in sheet.title :
            #項目数をカウント
            while val != '契約内容' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しくない項目数の場合は例外ファイルリストに追加
            if i != 72 :
                ary.append(path)
                print('\n' + path + '\n' + str(sheet.title) + str(i))
                break
        elif '脱毛' in sheet.title :
            #項目数をカウント
            while val != '医師から注意を受けている事や体質的に気になる事' and i < 10000 :
                val = sheet.cell(row = i, column = 2).value
                i = i + 1
            #正しくない項目数の場合は例外ファイルリストに追加
            if i != 62 :
                ary.append(path)
                print('\n' + path + '\n' + str(sheet.title) + str(i))
                break

#プログレスバーの表示
rate = bar_size
pro_bar =('=' * rate) + (' ' * (bar_size - rate))
print('\r' + '[' + pro_bar + ']' + ' ' + str(round(j / int(len_files) * 100, 2)) + '%', end='')

ary = list(set(ary))

#例外ファイルを要修正フォルダに移動させる
for one in ary :
    shutil.move(one, one.replace('/【A】/', '/【A】/【要修正】/'))

#例外ファイルリストをテキストファイルに出力
f = open('例外.txt', 'w', encoding = 'UTF-8')
for one in ary :
    f.write(one)
    f.write('\n')
f.close()

print('\nComplete!',end = '')